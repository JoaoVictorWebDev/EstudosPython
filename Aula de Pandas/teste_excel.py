import openpyxl
# Criar uma planilhar
book = openpyxl.Workbook()
# como visualizar as páginas existentes
print(book.sheetnames)
#como criar uma página
book.create_sheet('Funcionarios')
book.create_sheet('Frutas')

#como selecionar uma página
vagaodetrem = book['Frutas']
vagaodetrem.append(['Trem','Lote','Tipo de vagão'])
vagaodetrem.append(['VK-300','5','GONDOLA'])
#Página de Funcionarios
funcionarios = book['Funcionarios']

funcionarios.append(['Nome_Funcionarios','Idade' ,' Tempo de Serviço'])
funcionarios.remove(['João Victor','18 Anos','4 anos'])
funcionarios.append(['Carlos Gabriel','21 Anos','1 anos'])
funcionarios.append(['Matheus','22 Anos','4 anos'])
funcionarios.append(['Matheus','22 Anos','4 anos'])
funcionarios.append(['Matheus','22 Anos','4 anos'])
funcionarios.remove(['Matheus','22 Anos','4 anos'])
funcionarios.append(['Matheus','22 Anos','4 anos'])


#SALVAR  A PLANILHA
book.save('vagoes.xlsx')